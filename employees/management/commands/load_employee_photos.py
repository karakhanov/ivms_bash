"""
Загрузка фото сотрудников из Excel (встроенные изображения или папка с файлами).

Два режима:
1) Фото встроены в Excel (в ячейках) — не указывайте --photos-dir. По строке изображения
   определяется ПИНФЛ и фото сохраняется в БД по совпадению external_id = ПИНФЛ.
2) Фото в папке — укажите --photos-dir; имена файлов по ПИНФЛ (12345678901234.jpg) или
   колонка «фото» в Excel с именем файла.

Ожидания Excel: колонка ПИНФЛ (ПИНФЛ, pinfl, external_id и т.п.), при необходимости — фамилия.

Примеры:
  # Фото внутри .xlsx — сопоставление по ПИНФЛ со строкой картинки
  python manage.py load_employee_photos --excel /path/to/table.xlsx
  # Фото в папке
  python manage.py load_employee_photos --excel /path/to/employees.xlsx --photos-dir /path/to/photos
"""
import io
from pathlib import Path

import openpyxl
from django.core.files import File
from django.core.management.base import BaseCommand

from employees.models import Employee

# возможные названия колонки с идентификатором (ПИНФЛ)
EXTERNAL_ID_COLUMNS = ("external_id", "pinfl", "ПИНФЛ", "пинфл", "PINFL", "id", "ID")
# колонка с именем файла фото (если есть в Excel и используется --photos-dir)
PHOTO_COLUMNS = ("photo", "фото", "Photo", "Фото", "photo_path", "file")


def _extension_from_image_data(data: bytes) -> str:
    """Определение расширения по магическим байтам."""
    if data[:4] == b"\x89PNG":
        return ".png"
    if data[:2] == b"\xff\xd8":
        return ".jpg"
    if data[:4] == b"GIF8":
        return ".gif"
    return ".jpg"


class Command(BaseCommand):
    help = "Загрузить фото сотрудников из Excel и папки с изображениями (по ПИНФЛ/external_id)."

    def add_arguments(self, parser):
        parser.add_argument(
            "--excel",
            required=True,
            help="Путь к файлу Excel (.xlsx)",
        )
        parser.add_argument(
            "--photos-dir",
            default=None,
            help="Папка с фотографиями (если не указано — фото берутся из встроенных изображений Excel)",
        )
        parser.add_argument(
            "--column",
            default=None,
            help="Имя колонки с ПИНФЛ (по умолчанию ищем: external_id, pinfl, ПИНФЛ)",
        )
        parser.add_argument(
            "--sheet",
            type=int,
            default=0,
            help="Номер листа (0-based). По умолчанию первый.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Не сохранять, только показать, кого обновим",
        )

    def handle(self, *args, **options):
        excel_path = Path(options["excel"]).resolve()
        photos_dir = options.get("photos_dir")
        if photos_dir is not None:
            photos_dir = Path(photos_dir).resolve()
        column_name = options["column"]
        sheet_index = options["sheet"]
        dry_run = options["dry_run"]

        if not excel_path.exists():
            self.stderr.write(self.style.ERROR(f"Файл не найден: {excel_path}"))
            return
        if photos_dir is not None and not photos_dir.is_dir():
            self.stderr.write(self.style.ERROR(f"Папка не найдена: {photos_dir}"))
            return

        from_excel = photos_dir is None

        if from_excel:
            updated, skipped = self._load_from_embedded_images(
                excel_path, sheet_index, column_name, dry_run
            )
        else:
            updated, skipped = self._load_from_photos_dir(
                excel_path, photos_dir, sheet_index, column_name, dry_run
            )

        self.stdout.write(self.style.SUCCESS(f"Готово. Обновлено: {updated}, пропущено: {skipped}."))

    def _load_from_embedded_images(self, excel_path, sheet_index, column_name, dry_run):
        """Извлечь встроенные в лист изображения и сохранить по ПИНФЛ из той же строки."""
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if sheet_index >= len(wb.sheetnames):
            self.stderr.write(self.style.ERROR(f"Листа с индексом {sheet_index} нет."))
            return 0, 0
        ws = wb[wb.sheetnames[sheet_index]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            self.stderr.write(self.style.ERROR("В листе нет строк."))
            return 0, 0

        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        ext_id_col = self._find_pinfl_column(header, column_name)
        if ext_id_col is None:
            self.stderr.write(self.style.ERROR(f"Колонку ПИНФЛ не нашли. Заголовки: {header}"))
            return 0, 0

        images = getattr(ws, "_images", [])
        if not images:
            self.stderr.write(self.style.WARNING("В листе нет встроенных изображений (ws._images пуст)."))
            return 0, 0

        # Собираем (img, row, col) и сортируем по строке/столбцу — порядок как в листе
        image_rows = []
        for img in images:
            anchor = getattr(img, "anchor", None)
            if anchor is None:
                continue
            from_cell = getattr(anchor, "_from", None) or getattr(anchor, "from", None)
            if from_cell is None:
                continue
            row_num = getattr(from_cell, "row", 0)
            col_num = getattr(from_cell, "col", 0)
            if row_num is None:
                row_num = 0
            if col_num is None:
                col_num = 0
            image_rows.append((img, row_num, col_num))
        image_rows.sort(key=lambda x: (x[1], x[2]))

        updated = 0
        skipped = 0
        for img, row_num, _ in image_rows:
            # Якорь в вашем файле указывает на строку под картинкой — берём строку выше (сдвиг -1)
            data_row_idx = row_num
            if data_row_idx < 1 or data_row_idx >= len(rows):
                skipped += 1
                continue
            row = rows[data_row_idx]
            if not row or len(row) <= ext_id_col:
                skipped += 1
                continue
            external_id = row[ext_id_col]
            if external_id is None:
                skipped += 1
                continue
            external_id = str(external_id).strip()
            if not external_id:
                skipped += 1
                continue

            try:
                employee = Employee.objects.get(external_id=external_id)
            except Employee.DoesNotExist:
                self.stdout.write(self.style.WARNING(f"Сотрудник с ПИНФЛ {external_id} не найден в БД, пропуск."))
                skipped += 1
                continue

            try:
                data = img._data()
            except Exception as e:
                self.stdout.write(self.style.WARNING(f"Не удалось прочитать изображение для строки {data_row_idx + 1}: {e}"))
                skipped += 1
                continue
            if not data:
                skipped += 1
                continue

            ext = _extension_from_image_data(data)
            photo_name = f"{external_id}{ext}"

            if dry_run:
                self.stdout.write(f"[dry-run] Будет назначено фото: {employee} (ПИНФЛ {external_id})")
                updated += 1
                continue

            employee.photo.save(photo_name, File(io.BytesIO(data)), save=True)
            updated += 1
            self.stdout.write(f"Обновлено фото: {employee}")

        wb.close()
        return updated, skipped

    def _find_pinfl_column(self, header, column_name):
        if column_name:
            if column_name not in header:
                return None
            return header.index(column_name)
        for col_name in EXTERNAL_ID_COLUMNS:
            if col_name in header:
                return header.index(col_name)
        return None

    def _load_from_photos_dir(self, excel_path, photos_dir, sheet_index, column_name, dry_run):
        """Загрузить фото из папки, сопоставляя по ПИНФЛ из Excel."""
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        if sheet_index >= len(wb.sheetnames):
            self.stderr.write(self.style.ERROR(f"Листа с индексом {sheet_index} нет."))
            return 0, 0
        ws = wb[wb.sheetnames[sheet_index]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            self.stderr.write(self.style.ERROR("В листе нет строк."))
            return 0, 0

        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        ext_id_col = self._find_pinfl_column(header, column_name)
        if ext_id_col is None:
            self.stderr.write(self.style.ERROR(f"Колонку ПИНФЛ не нашли. Заголовки: {header}"))
            return 0, 0

        photo_col = None
        for p in PHOTO_COLUMNS:
            if p in header:
                photo_col = header.index(p)
                break

        updated = 0
        skipped = 0
        for row in rows[1:]:
            if not row or len(row) <= ext_id_col:
                continue
            external_id = row[ext_id_col]
            if external_id is None:
                continue
            external_id = str(external_id).strip()
            if not external_id:
                continue

            try:
                employee = Employee.objects.get(external_id=external_id)
            except Employee.DoesNotExist:
                self.stdout.write(self.style.WARNING(f"Сотрудник с external_id={external_id} не найден, пропуск."))
                skipped += 1
                continue

            if photo_col is not None and len(row) > photo_col and row[photo_col]:
                photo_name = str(row[photo_col]).strip()
            else:
                photo_name = None
            if not photo_name:
                for ext in ("jpg", "jpeg", "png", "JPG", "JPEG", "PNG"):
                    candidate = photos_dir / f"{external_id}.{ext}"
                    if candidate.exists():
                        photo_name = candidate.name
                        break
            if not photo_name:
                skipped += 1
                continue

            photo_path = photos_dir / photo_name
            if not photo_path.exists():
                self.stdout.write(self.style.WARNING(f"Файл не найден: {photo_path}"))
                skipped += 1
                continue

            if dry_run:
                self.stdout.write(f"[dry-run] Будет назначено: {employee} <- {photo_path}")
                updated += 1
                continue

            with open(photo_path, "rb") as f:
                employee.photo.save(photo_name, File(f), save=True)
            updated += 1
            self.stdout.write(f"Обновлено фото: {employee}")

        return updated, skipped
